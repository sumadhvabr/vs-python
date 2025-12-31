# ✅ Excel File Operations using openpyxl
import openpyxl

# -------------------------------
# Load an existing workbook
# -------------------------------
wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb.active

# -------------------------------
# Read a cell value
# -------------------------------
cell_value = sheet['A1'].value
print(f"Value in A1: {cell_value}")

# -------------------------------
# Write a value to a cell
# -------------------------------
sheet['B1'] = 42
wb.save('sample_modified.xlsx')

# -------------------------------
# Create a new workbook and add data
# -------------------------------
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet['A1'] = "Hello, Excel!"
new_sheet['B1'] = 100
new_wb.save('new_sample.xlsx')

# -------------------------------
# Read a data from a specific cell
# -------------------------------
print(f'The value stored in cell D5 is: {sheet["D5"].value}')

# -------------------------------
# Iterate through rows and columns
# -------------------------------
for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2):
    for cell in row:
        print(cell.value, end=' ')
    print()

# -------------------------------
# Get sheet names
# -------------------------------
sheet_names = wb.sheetnames
print(f"Sheet names in the workbook: {sheet_names}")

# -------------------------------
# Create a new sheet
# -------------------------------
new_sheet = wb.create_sheet(title="NewSheet")
wb.save('sample_modified.xlsx')

# -------------------------------
# Delete the created sheet
# -------------------------------
del wb['NewSheet']
wb.save('sample_modified.xlsx')

# -------------------------------
# Get max rows and columns
# -------------------------------
max_row = sheet.max_row
max_column = sheet.max_column
print(f"Max Row: {max_row}, Max Column: {max_column}")

# -------------------------------
# Read data from a range of cells
# -------------------------------
list1 = []
for value in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=4, values_only=True):
    list1.append(value)
print(list1)

# -------------------------------
# Merge and unmerge cells
# -------------------------------
sheet.merge_cells('C1:D1')
sheet['C1'] = "Merged Cell"
wb.save('sample_modified.xlsx')

sheet.unmerge_cells('C1:D1')
wb.save('sample_modified.xlsx')

# -------------------------------
# Display values in tabular format
# -------------------------------
for e1, e2, e3, e4 in list1:
    print("{:<10} {:<35} {:<10} {:<5}".format(e1, e2, e3, e4))

# -------------------------------
# Write to a new column (K)
# -------------------------------
sheet['K1'] = 'Total Sales'
sheet['K2'] = 1500
sheet['K3'] = 2000
sheet['K4'] = 2500
wb.save('sample_modified.xlsx')

# -------------------------------
# Add values in the 4 columns to get total sales
# -------------------------------
r_pos = 2
col_pos = 11  # Column K

t_sales = (
    sheet.cell(row=r_pos, column=col_pos).value +
    sheet.cell(row=r_pos + 1, column=col_pos).value +
    sheet.cell(row=r_pos + 2, column=col_pos).value
)

sheet.cell(row=r_pos + 3, column=col_pos).value = t_sales
wb.save('sample_modified.xlsx')

# -------------------------------
# Example: Loop to sum 4 regional sales (if data exists)
# -------------------------------
# Ensure that columns L, M, N, O have numeric data before running this block
for i in range(2, sheet.max_row + 1):
    N_sales = sheet.cell(row=i, column=11).value or 0
    S_sales = sheet.cell(row=i, column=12).value or 0
    E_sales = sheet.cell(row=i, column=13).value or 0
    W_sales = sheet.cell(row=i, column=14).value or 0
    total_sales = N_sales + S_sales + E_sales + W_sales
    sheet.cell(row=i, column=15).value = total_sales
wb.save('sample_modified.xlsx')

# -------------------------------
# Add a new row of data
# -------------------------------
n_r = (
    31, 'Cricket Premier League', 'PC', 2025, 'Sports', 'GameStudio', 2.60
)
sheet.append(n_r)
wb.save('VideoSales.xlsx')

# -------------------------------
# Delete the row which contains 'PC' in the 3rd column
# -------------------------------
rows_to_delete = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[2].value == 'PC':
        rows_to_delete.append(row[0].row)

for row_index in reversed(rows_to_delete):
    sheet.delete_rows(row_index)
wb.save('VideoSales.xlsx')

# -------------------------------
# Update the price of 'Cricket Premier League' to 3.00
# -------------------------------
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[1].value == 'Cricket Premier League':
        row[6].value = 3.00
wb.save('VideoSales.xlsx')

# -------------------------------
# Filter and display all games in the 'Sports' genre
# -------------------------------
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[4].value == 'Sports':
        print(f"Game: {row[1].value}, Platform: {row[2].value}, Price: {row[6].value}")

# -------------------------------
# Calculate average price
# -------------------------------
sheet['S1'] = 'Average Price'
total_price = 0
count = 0

for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    price = r[6].value
    if isinstance(price, (int, float)):
        total_price += price
        count += 1

if count > 0:
    average_price = total_price / count
    sheet['S2'] = average_price

wb.save('VideoSales.xlsx')

print("✅ All Excel operations completed successfully!")